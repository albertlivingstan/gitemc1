from tkinter import *
from tkinter.ttk import Combobox        
from tkinter import messagebox
import openpyxl
import pathlib

# Creating your window
root = Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False, False)
root.configure(bg="#326273")

# Check if the file exists, if not create it
file = pathlib.Path('Backend_data.xlsx')
if not file.exists():
    file = openpyxl.Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "PhoneNumber"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    file.save("Backend_data.xlsx")


# submit function for button 

def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = ageValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)

    # Load workbook and get the active sheet
    file = openpyxl.load_workbook('Backend_data.xlsx')
    sheet = file.active

    # Append data to the next available row
    next_row = sheet.max_row + 1
    sheet.cell(column=1, row=next_row, value=name)
    sheet.cell(column=2, row=next_row, value=contact)
    sheet.cell(column=3, row=next_row, value=age)
    sheet.cell(column=4, row=next_row, value=gender)
    sheet.cell(column=5, row=next_row, value=address)
    file.save('Backend_data.xlsx')

    # Show success message and clear entry fields
    messagebox.showinfo('Info', 'Detail added!')
    clear()

def clear():
    nameValue.set('')
    contactValue.set('')
    ageValue.set('')
    addressEntry.delete(1.0, END)


# this all code is used to creat window dsign

# Heading
Label(root, text="Please fill out this Entry form:", font="Arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

# Labels
Label(root, text='Name', font='Arial 12', bg="#326273", fg="#fff").place(x=50, y=100)
Label(root, text='Contact No.', font='Arial 12', bg="#326273", fg="#fff").place(x=50, y=150)
Label(root, text='Age', font='Arial 12', bg="#326273", fg="#fff").place(x=50, y=200)
Label(root, text='Gender', font='Arial 12', bg="#326273", fg="#fff").place(x=380, y=200)
Label(root, text='Address', font='Arial 12', bg="#326273", fg="#fff").place(x=50, y=250)

# Entry
nameValue = StringVar()
nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font='Arial 12')
nameEntry.place(x=200, y=100)

contactValue = StringVar()
contactEntry = Entry(root, textvariable=contactValue, width=45, bd=2, font='Arial 12')
contactEntry.place(x=200, y=150)

ageValue = StringVar()
ageEntry = Entry(root, textvariable=ageValue, width=10, bd=2, font='Arial 12')  # Reduced width for age entry
ageEntry.place(x=200, y=200)

gender_combobox = Combobox(root, values=['Male', 'Female'], font='Arial 12', state='readonly', width=14)
gender_combobox.place(x=440, y=200)
gender_combobox.set('Male')

addressEntry = Text(root, height=3, width=35, bd=2, font='Arial 12')
addressEntry.place(x=200, y=250)

# Buttons
Button(root, text="Submit", bg="#326273", fg="black", width=15, height=2, command=submit).place(x=200, y=350)
Button(root, text="Clear", bg="#326273", fg="black", width=15, height=2, command=clear).place(x=340, y=350)
Button(root, text="Exit", bg="#326273", fg="black", width=15, height=2, command=root.destroy).place(x=480, y=350)

root.mainloop()
